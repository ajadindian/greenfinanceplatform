from flask import Flask, request, jsonify, send_file, make_response
from werkzeug.utils import secure_filename
from flask_login import LoginManager, UserMixin
from tinydb import Query
from flask_cors import CORS
import os
from werkzeug.security import generate_password_hash, check_password_hash
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity, get_jwt
from Providers.OPENAILLMAPI import OPENAILLMAPI
import datetime
from database import db
import pandas as pd
import openpyxl
import logging
import numpy as np
from future import standard_library

import logger
standard_library.install_aliases()
from builtins import str
import io
from functools import wraps
from flask_jwt_extended import verify_jwt_in_request
from flask_cors import cross_origin
from supabase_manager import supabase_manager
import json
from Providers.OllamaLLMAPI import OllamaLLMAPI
from chart_tracking_agent import ChartTrackingAgent
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)
CORS(app, supports_credentials=True)
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY','abc123')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY','abc123')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = datetime.timedelta(minutes=30)
app.config['JWT_BLACKLIST_ENABLED'] = True
app.config['JWT_BLACKLIST_TOKEN_CHECKS'] = ['access']
jwt_blacklist = set()  # In-memory blacklist, consider using Redis for production
jwt = JWTManager(app) 

login_manager = LoginManager(app)
login_manager.login_view = 'login'

CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:3000", "http://localhost:5173"],  # Add your frontend URL
        "methods": ["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True
    }
})

app.config['JWT_TOKEN_LOCATION'] = ['headers']
app.config['JWT_HEADER_NAME'] = 'Authorization'
app.config['JWT_HEADER_TYPE'] = 'Bearer'

class Company:
    def __init__(self, company_data):
        self.id = company_data.get('id')
        self.name = company_data.get('name')
        self.email = company_data.get('email')
        self.password = company_data.get('password')

class User(UserMixin):
    def __init__(self, user_data):
        self.id = user_data.get('id')
        self.name = user_data.get('name')
        self.email = user_data.get('email')
        self.password = user_data.get('password')
        self.company_id = user_data.get('company_id')
        self.contact = user_data.get('contact')
        self.designation = user_data.get('designation')

    def get_id(self):
        return str(self.id)

@login_manager.user_loader
def load_user(user_id):
    User = Query()
    user_data = db.get_user(user_id)
    return User(user_data) if user_data else None   

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xlsx', 'xls'}


# Add this new model for projects
class Project:
    def __init__(self, project_data):
        self.id = project_data.get('id')
        self.name = project_data.get('name')
        self.structures = project_data.get('structures')
        self.rooms_per_structure = project_data.get('rooms_per_structure')
        self.user_id = project_data.get('user_id')

@app.route('/api/company/signup', methods=['POST'])
def company_signup():
    try:
        data = request.json
        logging.info(f"Received company signup request: {data}")

        if not data or 'name' not in data or 'email' not in data or 'password' not in data:
            logging.warning("Missing required fields in company signup request")
            return jsonify({"error": "Missing required fields"}), 400

        # Check for empty or whitespace-only company name
        if not data['name'] or not data['name'].strip():
            logging.warning("Empty company name provided")
            return jsonify({"error": "Company name cannot be empty"}), 400

        if db.company_exists_by_email(data['email']):
            logging.info(f"Company email already exists: {data['email']}")
            return jsonify({"error": "Company email already exists"}), 400

        if db.company_exists_by_name(data['name']):
            logging.info(f"Company name already exists: {data['name']}")
            return jsonify({"error": "Company name already exists"}), 400

        company_id = db.create_company({
            'name': data['name'].strip(),  # Remove leading/trailing whitespace
            'email': data['email'],
            'password': generate_password_hash(data['password'])
        })

        logging.info(f"Company created successfully with ID: {company_id}")
        return jsonify({"message": "Company created successfully", "id": company_id}), 201

    except Exception as e:
        logging.error(f"Error in company signup: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/company/login', methods=['POST'])
def company_login():
    data = request.json
    company_data = db.get_company_by_email(data['email'])
    
    if company_data and check_password_hash(company_data['password'], data['password']):
        access_token = create_access_token(identity=f"company_{company_data['id']}")
        return jsonify({"message": "Login successful", "token": access_token}), 200
    else:
        return jsonify({"error": "Invalid email or password"}), 401

@app.route('/api/user/signup', methods=['POST'])
def user_signup():
    data = request.json
    
    # Check if all required fields are present
    required_fields = ['name', 'email', 'password', 'company_id']
    missing_fields = [field for field in required_fields if field not in data]
    if missing_fields:
        return jsonify({"error": f"Missing required fields: {', '.join(missing_fields)}"}), 400

    if db.user_exists_by_email(data['email']):
        return jsonify({"error": "User email already exists"}), 400

    # Check if the company exists
    company = db.get_company(int(data['company_id']))
    if not company:
        return jsonify({"error": "Invalid company ID"}), 400

    user_id = db.create_user({
        'name': data['name'],
        'email': data['email'],
        'password': generate_password_hash(data['password']),
        'company_id': data['company_id'],
        'contact': data.get('contact'),
        'designation': data.get('designation')
    })

    return jsonify({"message": "User created successfully", "id": user_id}), 201

@app.route('/api/user/login', methods=['POST'])
def user_login():
    data = request.json
    user_data = db.get_user_by_email(data['email'])
    
    if user_data and check_password_hash(user_data['password'], data['password']):
        access_token = create_access_token(identity=f"user_{user_data['id']}")
        return jsonify({"message": "Login successful", "token": access_token}), 200
    else:
        return jsonify({"error": "Invalid email or password"}), 401

@app.route('/api/user', methods=['GET'])
@jwt_required()
def get_user():
    current_user_id = get_jwt_identity()
    if current_user_id.startswith('user_'):
        current_user_id = int(current_user_id.split('_')[1])
        user_data = db.get_user(current_user_id)
        if user_data:
            return jsonify({
                "name": user_data['name'],
                "email": user_data['email'],
                "company_id": user_data['company_id'],
                "designation": user_data['designation']
            }), 200
        else:
            return jsonify({"error": "User not found"}), 404
    else:
        return jsonify({"error": "Invalid token"}), 401

@app.route('/api/projects', methods=['GET'])
@jwt_required()
def get_projects():
    current_identity = get_jwt_identity()
    if current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        company_projects = db.get_company_projects(current_company_id)
        return jsonify([{"id": project.doc_id, "name": project['name']} for project in company_projects]), 200
    else:
        current_user_id = int(current_identity.split('_')[1])
        user_projects = db.get_user_projects(current_user_id)
        return jsonify([{"id": project.doc_id, "name": project['name']} for project in user_projects]), 200

@app.route('/api/projects', methods=['POST'])
@jwt_required()
def add_project():
    current_identity = get_jwt_identity()
    if not current_identity.startswith('company_'):
        return jsonify({"error": "Only companies can create projects"}), 403
    
    current_company_id = int(current_identity.split('_')[1])
    data = request.json
    
    # Check if project name already exists for this company
    Project = Query()
    existing_project = db.projects_db.get(
        (Project.name == data['name'].strip()) & 
        (Project.company_id == current_company_id)
    )
    
    if existing_project:
        return jsonify({"error": "A project with this name already exists"}), 400
    
    company_data = db.get_company(current_company_id)
    if not company_data:
        return jsonify({"error": "Company not found"}), 404
    
    # Sanitize company name and project name for file system
    company_name = secure_filename(company_data['name'].strip())
    project_name = secure_filename(data['name'].strip())
    
    # Create base upload folder if it doesn't exist
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    
    # Create company folder if it doesn't exist
    company_path = os.path.join(app.config['UPLOAD_FOLDER'], company_name)
    if not os.path.exists(company_path):
        os.makedirs(company_path)
    
    # Create project folder structure
    project_path = os.path.join(company_path, project_name)
    quotation_path = os.path.join(project_path, 'quotation')
    rag_cache_path = os.path.join(project_path, 'RAG_cache')
    
    try:
        os.makedirs(project_path, exist_ok=True)
        os.makedirs(quotation_path, exist_ok=True)
        os.makedirs(rag_cache_path, exist_ok=True)
        
        project_id = db.projects_db.insert({
            'name': data['name'].strip(),  # Store original name in database
            'company_id': current_company_id,
            'assigned_users': data.get('assigned_users', []),
            'path': project_path,
            'rag_cache_path': rag_cache_path
        })
        
        return jsonify({"message": "Project created successfully", "id": project_id}), 201
        
    except Exception as e:
        logging.error(f"Error creating project folders: {str(e)}")
        # Clean up any partially created folders if there was an error
        if os.path.exists(project_path):
            import shutil
            shutil.rmtree(project_path)
        return jsonify({"error": f"Failed to create project structure: {str(e)}"}), 500


def process_excel_to_text(file_path, is_quotation=False, project_name=""):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    chunks = []
    current_chunk = []
    current_length = 0
    max_chars = 8000  # Conservative estimate for ~8k tokens
    
    # Add header based on file type
    header = f"{'Quotation' if is_quotation else 'Actual'} File for Project: {project_name}"
    current_chunk.append(header)
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if not df.empty:
            sheet_header = f"\nSheet: {sheet_name}"
            columns_text = f"Columns: {', '.join(df.columns.tolist())}"
            
            # Start new chunk with headers
            if current_length > 0:
                chunks.append("\n".join(current_chunk))
                current_chunk = [header, sheet_header, columns_text]
                current_length = len(header) + len(sheet_header) + len(columns_text)
            else:
                current_chunk.extend([sheet_header, columns_text])
                current_length += len(sheet_header) + len(columns_text)
            
            # Process each row
            for idx, row in df.iterrows():
                row_text = []
                for col in df.columns:
                    value = row[col]
                    if pd.notna(value):
                        if isinstance(value, pd.Timestamp):
                            formatted_value = value.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(value, (np.integer, np.floating)):
                            formatted_value = f"{value:g}"
                        else:
                            formatted_value = str(value)
                        row_text.append(f"{col}: {formatted_value}")
                
                row_string = " | ".join(row_text)
                
                # If adding this row would exceed the limit, start a new chunk
                if current_length + len(row_string) > max_chars:
                    chunks.append("\n".join(current_chunk))
                    current_chunk = [header, sheet_header, row_string]
                    current_length = len(header) + len(sheet_header) + len(row_string)
                else:
                    current_chunk.append(row_string)
                    current_length += len(row_string)
    
    # Add the last chunk if it exists
    if current_chunk:
        chunks.append("\n".join(current_chunk))
    
    return chunks

# Initialize the chart tracking agent
chart_tracking_agent = ChartTrackingAgent()

@app.route('/api/upload', methods=['POST'])
@jwt_required()
def api_upload():
    current_identity = get_jwt_identity()
    logging.info(f"File upload initiated by: {current_identity}")
    if 'file' not in request.files:
        logging.warning("No file part in the request")
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        logging.warning("No selected file")
        return jsonify({"error": "No selected file"}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        project_id = request.form.get('project')
        is_quotation = request.form.get('is_quotation', 'false').lower() == 'true'
        is_update = request.form.get('is_update', 'false').lower() == 'true'
        
        logging.info(f"Processing file: {filename} for project: {project_id}, is_quotation: {is_quotation}")
        
        if not project_id:
            logging.warning("No project specified")
            return jsonify({"error": "No project specified"}), 400
        
        project = db.get_project(int(project_id))
        if not project:
            logging.warning(f"Project not found: {project_id}")
            return jsonify({"error": "Project not found"}), 404
        
        # Check if the current user has access to the project
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                return jsonify({"error": "Unauthorized access to project"}), 403
        else:
            return jsonify({"error": "Invalid token"}), 401
        
        project_path = project['path']
        
        # Check if this file already exists
        existing_files = db.get_project_files(int(project_id))
        existing_file = next((f for f in existing_files if f['name'] == filename), None)
        
        # Determine the save path
        if is_quotation:
            rel_path = 'quotation'
            path = os.path.join(project_path, 'quotation')
        else:
            now = datetime.datetime.now()
            date_folder = now.strftime('%Y-%m-%d')
            time_str = now.strftime('%H-%M-%S')
            rel_path = date_folder
            path = os.path.join(project_path, date_folder)
            
            # If this is an update, use the existing file's path
            if is_update and existing_file:
                rel_path = existing_file['relative_path']
                path = os.path.join(project_path, rel_path)
                # Delete the old file
                if os.path.exists(existing_file['path']):
                    os.remove(existing_file['path'])
            else:
                filename = f"{time_str}_{filename}"
        
        os.makedirs(path, exist_ok=True)
        file_path = os.path.join(path, filename)
        
        try:
            # First save the file locally
            file.save(file_path)
            logging.info(f"File saved to: {file_path}")
            
            # Process and add to Supabase
            if file.filename.endswith(('.xlsx', '.xls')):
                text_chunks = process_excel_to_text(
                    file_path,
                    is_quotation=is_quotation,
                    project_name=project['name']
                )
                
                # Clear any existing documents for this file if it's an update
                if is_update:
                    try:
                        supabase_manager.delete_file(int(project_id), file_path)
                        logging.info(f"Deleted existing file data from Supabase for update")
                    except Exception as e:
                        logging.error(f"Error deleting existing file data: {str(e)}")
                
                # Add new chunks to Supabase
                for i, chunk in enumerate(text_chunks):
                    metadata = {
                        "file_name": filename,
                        "file_type": "excel",
                        "is_quotation": is_quotation,
                        "file_path": file_path,
                        "chunk_index": i,
                        "total_chunks": len(text_chunks)
                    }
                    supabase_manager.add_document(int(project_id), chunk, metadata)
                
                # Save file metadata
                file_info = {
                    "name": filename,
                    "relative_path": rel_path,
                    "addedBy": "System",
                    "dateAdded": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "isUpdate": is_update
                }
                db.save_file_metadata(int(project_id), file_info)
                
                # Now that Supabase is updated, trigger chart updates
                try:
                    logging.info(f"Initiating chart updates for project {project_id}")
                    chart_tracking_agent.update_project_charts(int(project_id))
                    logging.info("Charts updated successfully")
                except Exception as chart_error:
                    logging.error(f"Error updating charts: {str(chart_error)}")
                
                return jsonify({"message": "File uploaded and processed successfully"}), 200
                
            else:
                with open(file_path, 'r') as f:
                    file_content = f.read()
                
                # Clear any existing documents for this file if it's an update
                if is_update:
                    try:
                        supabase_manager.delete_file(int(project_id), file_path)
                        logging.info(f"Deleted existing file data from Supabase for update")
                    except Exception as e:
                        logging.error(f"Error deleting existing file data: {str(e)}")
                
                metadata = {
                    "file_name": filename,
                    "file_type": "text",
                    "is_quotation": is_quotation,
                    "file_path": file_path
                }
                supabase_manager.add_document(int(project_id), file_content, metadata)
                
                # Save file metadata
                file_info = {
                    "name": filename,
                    "relative_path": rel_path,
                    "addedBy": "System",
                    "dateAdded": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    "isUpdate": is_update
                }
                db.save_file_metadata(int(project_id), file_info)
                
                # Now that Supabase is updated, trigger chart updates
                try:
                    logging.info(f"Initiating chart updates for project {project_id}")
                    chart_tracking_agent.update_project_charts(int(project_id))
                    logging.info("Charts updated successfully")
                except Exception as chart_error:
                    logging.error(f"Error updating charts: {str(chart_error)}")
                
                return jsonify({"message": "File uploaded and processed successfully"}), 200
                
        except Exception as e:
            logging.error(f"Error processing file: {str(e)}")
            return jsonify({"error": "File processing failed", "details": str(e)}), 500
    
    logging.warning(f"File type not allowed: {file.filename}")
    return jsonify({"error": "File type not allowed"}), 400

@app.route('/api/chat', methods=['POST'])
@jwt_required()
def chat():
    current_identity = get_jwt_identity()
    data = request.json
    query = data.get('query')
    project_id = data.get('project')
    
    logging.info(f"\n\nChat request received for project: {project_id}")
    logging.info(f"\n\nQuery: {query}")

    if not query or not project_id:
        logging.warning("\n\nMissing query or project in chat request")
        return jsonify({"error": "Missing query or project"}), 400
    
    if current_identity.startswith('user_'):
        current_user_id = int(current_identity.split('_')[1])
        user_data = db.get_user(current_user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        project = db.get_project(int(project_id))
        if not project or project['company_id'] != user_data['company_id']:
            return jsonify({"error": "Project not found or unauthorized"}), 404
    elif current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        project = db.get_project(int(project_id))
        if not project or project['company_id'] != current_company_id:
            return jsonify({"error": "Project not found or unauthorized"}), 404
    else:
        return jsonify({"error": "Invalid token"}), 401
    
    try:
        # Use Supabase for vector search
        results = supabase_manager.query(int(project_id), query)
        
        # Combine relevant content
        relevant_info = "\n".join([result['content'] for result in results])
        
        # Use OpenAI for final response
        openai_llm = OPENAILLMAPI()
        final_response = openai_llm.get_ai_response(query, relevant_info)
        
        return jsonify({"response": final_response}), 200
        
    except Exception as e:
        logging.error(f"Error in chat endpoint: {str(e)}")
        return jsonify({"error": "An error occurred processing your request"}), 500

@app.route('/api/companies', methods=['GET'])
def get_companies():
    companies = db.get_all_companies()
    return jsonify([
        {
            "id": company.doc_id,
            "name": company['name']
        } for company in companies
    ]), 200

def process_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheets_data = {}

    for sheet_name in workbook.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        if not df.empty:
            # Get all column names and create column nodes
            columns = df.columns.tolist()
            records = []
            
            # Create column relationship metadata
            column_relationships = {}
            for col1 in columns:
                for col2 in columns:
                    if col1 != col2:
                        correlation = df[col1].corr(df[col2]) if pd.api.types.is_numeric_dtype(df[col1]) and pd.api.types.is_numeric_dtype(df[col2]) else None
                        column_relationships[f"{col1}_{col2}"] = {
                            "source": col1,
                            "target": col2,
                            "type": "column_relation",
                            "correlation": correlation if correlation and not np.isnan(correlation) else None,
                            "relationship_type": "numerical" if correlation is not None else "categorical"
                        }
            
            # Process each row while preserving column relationships
            for _, row in df.iterrows():
                processed_record = {}
                value_relationships = {}
                
                for col in columns:
                    value = row[col]
                    # Handle different data types
                    if isinstance(value, pd.Timestamp):
                        processed_value = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, np.integer):
                        processed_value = int(value)
                    elif isinstance(value, np.floating):
                        processed_value = float(value)
                    else:
                        processed_value = str(value) if pd.notna(value) else None
                        
                    if processed_value is not None:
                        processed_record[col] = processed_value
                        
                        # Create value relationships between non-null values
                        for other_col in columns:
                            if other_col != col and pd.notna(row[other_col]):
                                other_value = row[other_col]
                                if isinstance(other_value, pd.Timestamp):
                                    other_value = other_value.strftime('%Y-%m-%d %H:%M:%S')
                                elif isinstance(other_value, (np.integer, np.floating)):
                                    other_value = float(other_value)
                                else:
                                    other_value = str(other_value)
                                
                                rel_key = f"{col}_{other_col}_value"
                                value_relationships[rel_key] = {
                                    "source": str(processed_value),
                                    "target": str(other_value),
                                    "source_column": col,
                                    "target_column": other_col,
                                    "type": "value_relation"
                                }
                
                records.append({
                    "data": processed_record,
                    "value_relationships": value_relationships,
                    "column_relationships": column_relationships
                })
            
            sheets_data[sheet_name] = {
                "records": records,
                "metadata": {
                    "columns": columns,
                    "column_relationships": column_relationships,
                    "total_rows": len(records)
                }
            }
        else:
            logging.warning(f"Sheet {sheet_name} in {file_path} is empty. Skipping.")

    return sheets_data

# Add this function to create the upload folder if it doesn't exist
# def create_upload_folder():
#     upload_folder = app.config['UPLOAD_FOLDER']
#     if not os.path.exists(upload_folder):
#         os.makedirs(upload_folder)

@app.route('/api/debug/token', methods=['GET'])
@jwt_required()
def debug_token():
    current_identity = get_jwt_identity()
    return jsonify({"message": "Token is valid", "identity": current_identity}), 200

def cors_preflight():
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if request.method == 'OPTIONS':
                response = make_response()
                response.headers.add('Access-Control-Allow-Origin', 'http://localhost:5173')  # Or specify your frontend URL
                response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
                response.headers.add('Access-Control-Allow-Methods', 'GET,POST,PUT,DELETE,OPTIONS')
                response.headers.add('Access-Control-Allow-Credentials', 'true')
                return response
            return f(*args, **kwargs)
        return wrapped
    return decorator

def jwt_or_options():
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if request.method == 'OPTIONS':
                return cors_preflight()(lambda *a, **kw: None)(*args, **kwargs)
            try:
                verify_jwt_in_request()
            except Exception as e:
                return jsonify({"error": "Invalid or missing token"}), 401
            return f(*args, **kwargs)
        return wrapped
    return decorator

@app.route('/api/projects/<int:project_id>', methods=['GET', 'OPTIONS'])
@cross_origin(supports_credentials=True)
@jwt_required()  # Add this decorator
def get_project_details(project_id):
    if request.method == 'OPTIONS':
        return '', 200
    current_identity = get_jwt_identity()
    project = db.get_project(project_id)
    
    if not project:
        return jsonify({"error": "Project not found"}), 404
    
    # Check authorization
    if current_identity.startswith('user_'):
        current_user_id = int(current_identity.split('_')[1])
        if current_user_id not in project.get('assigned_users', []):
            return jsonify({"error": "Unauthorized access to project"}), 403
    elif current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        if project['company_id'] != current_company_id:
            return jsonify({"error": "Unauthorized access to project"}), 403
    
    # Get project files
    project_files = db.get_project_files(project_id)
    
    return jsonify({
        "id": project_id,
        "name": project['name'],
        "description": project.get('description', ''),
        "files": project_files
    }), 200

@app.route('/api/projects/<int:project_id>/pl', methods=['GET', 'OPTIONS'])
@cross_origin(supports_credentials=True)
@jwt_required()  # Add this decorator
def get_project_pl(project_id):
    if request.method == 'OPTIONS':
        return '', 200
    current_identity = get_jwt_identity()
    project = db.get_project(project_id)
    
    if not project:
        return jsonify({"error": "Project not found"}), 404
    
    # Add authorization check similar to get_project_details
    if current_identity.startswith('user_'):
        current_user_id = int(current_identity.split('_')[1])
        if current_user_id not in project.get('assigned_users', []):
            return jsonify({"error": "Unauthorized access to project"}), 403
    elif current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        if project['company_id'] != current_company_id:
            return jsonify({"error": "Unauthorized access to project"}), 403
    
    # This is a placeholder. Implement your actual P/L calculation logic
    return jsonify({
        "totalRevenue": 100000,
        "totalCost": 80000,
        "netProfit": 20000
    }), 200

@app.route('/api/projects/<int:project_id>/files', methods=['GET', 'OPTIONS'])
@cross_origin(supports_credentials=True)
@jwt_required()
def get_project_files(project_id):
    if request.method == 'OPTIONS':
        return '', 200
    current_identity = get_jwt_identity()
    project = db.get_project(project_id)
    
    if not project:
        return jsonify({"error": "Project not found"}), 404
    
    # Add authorization check
    if current_identity.startswith('user_'):
        current_user_id = int(current_identity.split('_')[1])
        if current_user_id not in project.get('assigned_users', []):
            return jsonify({"error": "Unauthorized access to project"}), 403
    elif current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        if project['company_id'] != current_company_id:
            return jsonify({"error": "Unauthorized access to project"}), 403
    
    project_files = db.get_project_files(project_id)
    
    # Filter out the consolidated data text file
    filtered_files = [
        file for file in project_files 
        if not (file.get('name') == 'consolidated_data.txt' and 'text_content' in file.get('relative_path', ''))
    ]
    
    return jsonify(filtered_files), 200

@app.route('/api/projects/<int:project_id>/files/<int:file_id>', methods=['DELETE', 'OPTIONS'])
@cross_origin(supports_credentials=True)
@jwt_required()
def delete_project_file(project_id, file_id):
    try:
        # Get file info first
        file_info = db.get_project_file(project_id, file_id)
        if not file_info:
            return jsonify({"error": "File not found"}), 404
            
        # Delete from Supabase - pass the file path directly
        success = supabase_manager.delete_file(project_id, file_info['path'])
        
        if success:
            # Delete the local file if it exists
            if os.path.exists(file_info['path']):
                os.remove(file_info['path'])
            return jsonify({"message": "File deleted successfully"}), 200
        else:
            return jsonify({"error": "Failed to delete file"}), 500
            
    except Exception as e:
        logging.error(f"Error in delete_project_file: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/projects/<int:project_id>/files/<int:file_id>/download', methods=['GET', 'OPTIONS'])
@cors_preflight()
@jwt_or_options()
def download_project_file(project_id, file_id):
    if request.method == 'OPTIONS':
        return '', 200
    current_identity = get_jwt_identity()
    project = db.get_project(project_id)
    
    if not project:
        return jsonify({"error": "Project not found"}), 404
    
    # Check authorization
    if current_identity.startswith('user_'):
        current_user_id = int(current_identity.split('_')[1])
        if current_user_id not in project.get('assigned_users', []):
            return jsonify({"error": "Unauthorized access to project"}), 403
    elif current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        if project['company_id'] != current_company_id:
            return jsonify({"error": "Unauthorized access to project"}), 403
    
    file_data = db.get_project_file(project_id, file_id)
    if file_data:
        return send_file(
            io.BytesIO(file_data['content']),
            mimetype=file_data['mime_type'],
            as_attachment=True,
            download_name=file_data['name']
        )
    else:
        return jsonify({"error": "File not found"}), 404

@app.route('/api/predict', methods=['POST'])
@jwt_required()
def predict():
    current_identity = get_jwt_identity()
    data = request.json
    query = data.get('query')
    project_id = data.get('project')
    
    logging.info(f"\n\nPrediction request received for project: {project_id}")
    logging.info(f"\n\nQuery: {query}")

    if not query or not project_id:
        logging.warning("\n\nMissing query or project in prediction request")
        return jsonify({"error": "Missing query or project"}), 400
    
    # Authorization check
    if current_identity.startswith('user_'):
        current_user_id = int(current_identity.split('_')[1])
        user_data = db.get_user(current_user_id)
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        project = db.get_project(int(project_id))
        if not project or project['company_id'] != user_data['company_id']:
            return jsonify({"error": "Project not found or unauthorized"}), 404
    elif current_identity.startswith('company_'):
        current_company_id = int(current_identity.split('_')[1])
        project = db.get_project(int(project_id))
        if not project or project['company_id'] != current_company_id:
            return jsonify({"error": "Project not found or unauthorized"}), 404
    else:
        return jsonify({"error": "Invalid token"}), 401
    
    try:
        # Use Supabase for vector search
        results = supabase_manager.query(int(project_id), query)
        
        # Combine relevant content
        relevant_info = "\n".join([result['content'] for result in results])
        
        # Use OpenAI for prediction
        openai_llm = OPENAILLMAPI()
        prediction_response = openai_llm.get_prediction(query, relevant_info)
        
        return jsonify(json.loads(prediction_response)), 200
        
    except Exception as e:
        logging.error(f"Error in prediction endpoint: {str(e)}")
        return jsonify({"error": "An error occurred processing your request"}), 500

@app.route('/api/projects/<int:project_id>/charts', methods=['POST'])
@jwt_required()
def save_project_chart(project_id):
    try:
        current_identity = get_jwt_identity()
        data = request.json

        # Validate required fields
        if not all(k in data for k in ['name', 'query', 'chart_data']):
            return jsonify({"error": "Missing required fields"}), 400

        # Add user information
        if current_identity.startswith('user_'):
            data['created_by'] = int(current_identity.split('_')[1])
        elif current_identity.startswith('company_'):
            data['created_by'] = int(current_identity.split('_')[1])

        # Save the chart
        chart_id = db.save_chart(project_id, data)
        if chart_id:
            return jsonify({
                "message": "Chart saved successfully",
                "id": chart_id
            }), 201
        else:
            return jsonify({"error": "Failed to save chart"}), 500

    except Exception as e:
        logging.error(f"Error saving chart: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/charts', methods=['GET'])
@jwt_required()
def get_project_charts(project_id):
    try:
        charts = db.get_project_charts(project_id)
        return jsonify([{
            "id": chart.doc_id,
            "name": chart['name'],
            "query": chart['query'],
            "chart_data": chart['chart_data'],
            "created_at": chart['created_at'],
            "created_by": chart.get('created_by')
        } for chart in charts]), 200

    except Exception as e:
        logging.error(f"Error retrieving charts: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/charts/<int:chart_id>', methods=['DELETE'])
@jwt_required()
def delete_project_chart(project_id, chart_id):
    try:
        if db.delete_chart(project_id, chart_id):
            return jsonify({"success": True}), 200
        else:
            return jsonify({"error": "Failed to delete chart"}), 500

    except Exception as e:
        logging.error(f"Error deleting chart: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/dashboard/charts/<int:chart_id>', methods=['POST'])
@jwt_required()
def add_chart_to_dashboard(project_id, chart_id):
    try:
        # Verify project and chart exist
        project = db.get_project(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        chart = db.charts_db.get(doc_id=chart_id)
        if not chart or chart['project_id'] != project_id:
            return jsonify({"error": "Chart not found or doesn't belong to project"}), 404

        # Pin the chart
        if db.pin_chart(project_id, chart_id):
            return jsonify({"message": "Chart added to dashboard successfully"}), 200
        else:
            return jsonify({"error": "Failed to add chart to dashboard"}), 500

    except Exception as e:
        logging.error(f"Error adding chart to dashboard: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/dashboard/charts', methods=['GET'])
@jwt_required()
def get_dashboard_charts(project_id):
    try:
        # Verify project exists
        project = db.get_project(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Get pinned charts
        pinned_charts = db.get_pinned_charts(project_id)
        return jsonify([{
            "id": chart.doc_id,
            "name": chart['name'],
            "query": chart['query'],
            "chart_data": chart['chart_data'],
            "created_at": chart['created_at'],
            "created_by": chart.get('created_by')
        } for chart in pinned_charts]), 200

    except Exception as e:
        logging.error(f"Error retrieving dashboard charts: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/dashboard/charts/<int:chart_id>', methods=['DELETE'])
@jwt_required()
def remove_chart_from_dashboard(project_id, chart_id):
    try:
        # Verify project and chart exist
        project = db.get_project(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        chart = db.charts_db.get(doc_id=chart_id)
        if not chart or chart['project_id'] != project_id:
            return jsonify({"error": "Chart not found or doesn't belong to project"}), 404

        # Unpin the chart
        if db.unpin_chart(project_id, chart_id):
            return jsonify({"message": "Chart removed from dashboard successfully"}), 200
        else:
            return jsonify({"error": "Failed to remove chart from dashboard"}), 500

    except Exception as e:
        logging.error(f"Error removing chart from dashboard: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/dashboard/layouts', methods=['POST'])
@jwt_required()
def save_dashboard_layout(project_id):
    try:
        current_identity = get_jwt_identity()
        data = request.json
        
        # Log the received data for debugging
        logging.info(f"Received layout data: {json.dumps(data, indent=2)}")

        # Validate required fields exist
        if not data:
            return jsonify({"error": "No data provided"}), 400
            
        if 'name' not in data:
            return jsonify({"error": "Layout name is required"}), 400
            
        if not data.get('name').strip():
            return jsonify({"error": "Layout name cannot be empty"}), 400

        # Initialize empty structures if not provided
        data['layout_data'] = data.get('layout_data', {})
        data['charts'] = data.get('charts', [])

        # Verify project access
        project = db.get_project(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Check authorization
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                return jsonify({"error": "Unauthorized access to project"}), 403

        # Save the layout
        layout_id = db.save_dashboard_layout(project_id, data)
        if layout_id:
            return jsonify({
                "message": "Dashboard layout saved successfully",
                "id": layout_id
            }), 201
        else:
            return jsonify({"error": "Failed to save dashboard layout"}), 500

    except Exception as e:
        logging.error(f"Error saving dashboard layout: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/projects/<int:project_id>/dashboard/layouts', methods=['GET'])
@jwt_required()
def get_dashboard_layouts(project_id):
    try:
        # Verify project access and authorization
        current_identity = get_jwt_identity()
        project = db.get_project(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Check authorization
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                return jsonify({"error": "Unauthorized access to project"}), 403

        layouts = db.get_dashboard_layouts(project_id)
        return jsonify(layouts), 200

    except Exception as e:
        logging.error(f"Error retrieving dashboard layouts: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/dashboard/layouts/<int:layout_id>', methods=['GET'])
@jwt_required()
def get_dashboard_layout(project_id, layout_id):
    try:
        # Verify project access and authorization
        current_identity = get_jwt_identity()
        project = db.get_project(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Check authorization
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                return jsonify({"error": "Unauthorized access to project"}), 403

        layout = db.get_dashboard_layout(project_id, layout_id)
        if layout:
            return jsonify(layout), 200
        else:
            return jsonify({"error": "Layout not found"}), 404

    except Exception as e:
        logging.error(f"Error retrieving dashboard layout: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/user/settings', methods=['GET', 'OPTIONS'])
@cross_origin(supports_credentials=True, origins=["http://localhost:3000", "http://localhost:5173"])
@jwt_required()
def get_settings():
    if request.method == 'OPTIONS':
        return '', 200
        
    try:
        current_identity = get_jwt_identity()
        if not current_identity:
            return jsonify({"error": "Unauthorized"}), 401
            
        # Handle both company and user tokens
        try:
            if current_identity.startswith('company_'):
                entity_id = int(current_identity.split('_')[1])
                entity = db.get_company(entity_id)
                entity_type = 'company'
            else:  # user token
                entity_id = int(current_identity.split('_')[1])
                entity = db.get_user(entity_id)
                entity_type = 'user'
        except (IndexError, ValueError):
            logging.error(f"Invalid identity format: {current_identity}")
            return jsonify({"error": "Invalid token format"}), 400
        
        if not entity:
            logging.error(f"{entity_type.capitalize()} not found: {entity_id}")
            return jsonify({"error": f"{entity_type.capitalize()} not found"}), 404
            
        # Get or create settings
        settings = entity.get('settings', {})
        if not settings:
            settings = {
                'display_name': entity.get('name', ''),
                'notifications_enabled': True,
                'theme_preference': 'light'
            }
            # Update the entity with default settings
            if entity_type == 'company':
                db.companies_db.update({'settings': settings}, doc_ids=[entity_id])
            else:
                db.users_db.update({'settings': settings}, doc_ids=[entity_id])
        
        return jsonify({
            'email': entity.get('email'),
            'display_name': settings.get('display_name', entity.get('name', '')),
            'notifications_enabled': settings.get('notifications_enabled', True),
            'theme_preference': settings.get('theme_preference', 'light')
        }), 200
        
    except Exception as e:
        logging.error(f"Error in get_settings: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/user/logout', methods=['POST'])
@jwt_required()
def logout():
    try:
        jti = get_jwt()['jti']
        jwt_blacklist.add(jti)
        return jsonify({"message": "Successfully logged out"}), 200
    except Exception as e:
        logging.error(f"Error in logout: {str(e)}")
        return jsonify({"error": "Logout failed"}), 500

# Add these endpoints to handle payment-related requests temporarily
@app.route('/api/user/payment', methods=['GET', 'OPTIONS'])
@cross_origin(supports_credentials=True, origins=["http://localhost:3000", "http://localhost:5173"])
@jwt_required()
def get_payment_info():
    if request.method == 'OPTIONS':
        return '', 200
        
    # Temporary placeholder response
    return jsonify({
        "message": "Payment functionality coming soon",
        "last4": "****",
        "expiry_month": None,
        "expiry_year": None,
        "brand": None
    }), 200

@app.route('/api/user/payment', methods=['POST', 'OPTIONS'])
@cross_origin(supports_credentials=True, origins=["http://localhost:3000", "http://localhost:5173"])
@jwt_required()
def update_payment_info():
    if request.method == 'OPTIONS':
        return '', 200
        
    return jsonify({
        "message": "Payment updates not yet implemented"
    }), 501  # 501 Not Implemented

# Add a migration function to ensure all existing users have settings
def migrate_existing_users():
    try:
        users = db.users_db.all()
        for user in users:
            user_id = user.doc_id
            if 'settings' not in user:
                default_settings = {
                    'display_name': user.get('name', ''),
                    'notifications_enabled': True,
                    'theme_preference': 'light'
                }
                db.update_user_settings(user_id, default_settings)
                logger.info(f"Migrated settings for user {user_id}")
    except Exception as e:
        logger.error(f"Error migrating user settings: {str(e)}")

@app.route('/api/user/settings/password', methods=['POST', 'PATCH', 'OPTIONS'])
@cross_origin(supports_credentials=True, origins=["http://localhost:3000", "http://localhost:5173"])
@jwt_required()
def update_password():
    if request.method == 'OPTIONS':
        return '', 200
        
    try:
        current_identity = get_jwt_identity()
        logging.info(f"Current identity: {current_identity}")
        
        # Log the raw request data
        raw_data = request.get_json()
        logging.info(f"Raw request data: {raw_data}")
        
        if not raw_data:
            return jsonify({"error": "No data provided"}), 400
            
        # Check the exact keys being sent
        logging.info(f"Request keys: {raw_data.keys()}")
        
        # Check if we're getting the password fields with different key names
        current_password = raw_data.get('currentPassword') or raw_data.get('current_password') or raw_data.get('oldPassword')
        new_password = raw_data.get('newPassword') or raw_data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({
                "error": "Missing password fields",
                "received_keys": list(raw_data.keys()),
                "required_keys": ["currentPassword", "newPassword"]
            }), 400
            
        try:
            if current_identity.startswith('company_'):
                entity_id = int(current_identity.split('_')[1])
                company = db.get_company(entity_id)
                
                if not company:
                    return jsonify({"error": "Company not found"}), 404
                
                # Log the comparison values (be careful with this in production!)
                logging.info(f"Company found: {company.get('name')}")
                logging.info(f"Stored hash type: {type(company['password'])}")
                logging.info(f"Input password type: {type(current_password)}")
                
                # Try the verification
                try:
                    is_valid = check_password_hash(company['password'], current_password)
                    logging.info(f"Password verification result: {is_valid}")
                except Exception as verify_error:
                    logging.error(f"Error during password verification: {str(verify_error)}")
                    return jsonify({"error": "Password verification failed"}), 400
                
                if is_valid:
                    # Update password
                    new_hash = generate_password_hash(new_password)
                    db.companies_db.update({
                        'password': new_hash
                    }, doc_ids=[entity_id])
                    return jsonify({"message": "Password updated successfully"}), 200
                else:
                    return jsonify({
                        "error": "Current password is incorrect",
                        "debug_info": "Password verification failed"
                    }), 400
                    
            else:
                # Handle user case...
                return jsonify({"error": "User password updates not implemented"}), 501
                
        except (IndexError, ValueError) as e:
            logging.error(f"Error processing request: {str(e)}")
            return jsonify({"error": f"Invalid token format: {str(e)}"}), 400
            
    except Exception as e:
        logging.error(f"Error in update_password: {str(e)}")
        return jsonify({
            "error": "Internal server error",
            "details": str(e)
        }), 500

@app.route('/api/projects/<int:project_id>/prompts', methods=['POST'])
@jwt_required()
def save_project_prompt(project_id):
    try:
        current_identity = get_jwt_identity()
        data = request.json
        logging.info(f"Received prompt save request with data: {json.dumps(data, indent=2)}")
        
        # Validate request data exists
        if not data:
            logging.error("No JSON data received in request")
            return jsonify({"error": "No data provided"}), 400

        # Validate required fields - only content is required now
        if not data.get('content'):
            logging.error("Missing content field")
            return jsonify({"error": "Content is required"}), 400

        # Verify project exists
        project = db.get_project(project_id)
        if not project:
            logging.error(f"Project not found: {project_id}")
            return jsonify({"error": "Project not found"}), 404

        # Check authorization
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                logging.error(f"User {current_user_id} not authorized for project {project_id}")
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                logging.error(f"Company {current_company_id} not authorized for project {project_id}")
                return jsonify({"error": "Unauthorized access to project"}), 403

        # Prepare prompt data
        prompt_data = {
            'content': data['content'].strip(),
            'tags': [tag.strip() for tag in data.get('tags', []) if isinstance(tag, str)],
            'created_by': current_identity
        }

        # Save the prompt
        prompt_id = db.save_prompt(project_id, prompt_data)
        
        if prompt_id:
            logging.info(f"Successfully saved prompt with ID: {prompt_id}")
            return jsonify({
                "message": "Prompt saved successfully",
                "id": prompt_id
            }), 201
        else:
            logging.error("Failed to save prompt in database")
            return jsonify({"error": "Failed to save prompt"}), 500

    except Exception as e:
        logging.error(f"Unexpected error saving prompt: {str(e)}", exc_info=True)
        return jsonify({
            "error": "An unexpected error occurred",
            "details": str(e)
        }), 500

@app.route('/api/projects/<int:project_id>/prompts', methods=['GET'])
@jwt_required()
def get_project_prompts(project_id):
    try:
        current_identity = get_jwt_identity()
        
        # Verify project exists
        project = db.get_project(project_id)
        if not project:
            logging.error(f"Project not found: {project_id}")
            return jsonify({"error": "Project not found"}), 404

        # Check authorization
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                return jsonify({"error": "Unauthorized access to project"}), 403

        prompts = db.get_project_prompts(project_id)
        return jsonify(prompts), 200

    except Exception as e:
        logging.error(f"Error retrieving prompts: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

@app.route('/api/projects/<int:project_id>/prompts/<int:prompt_id>', methods=['DELETE'])
@jwt_required()
def delete_project_prompt(project_id, prompt_id):
    try:
        current_identity = get_jwt_identity()
        
        # Verify project exists
        project = db.get_project(project_id)
        if not project:
            logging.error(f"Project not found: {project_id}")
            return jsonify({"error": "Project not found"}), 404

        # Check authorization
        if current_identity.startswith('user_'):
            current_user_id = int(current_identity.split('_')[1])
            if current_user_id not in project.get('assigned_users', []):
                return jsonify({"error": "Unauthorized access to project"}), 403
        elif current_identity.startswith('company_'):
            current_company_id = int(current_identity.split('_')[1])
            if project['company_id'] != current_company_id:
                return jsonify({"error": "Unauthorized access to project"}), 403

        if db.delete_prompt(project_id, prompt_id):
            return jsonify({"message": "Prompt deleted successfully"}), 200
        else:
            return jsonify({"error": "Failed to delete prompt"}), 500

    except Exception as e:
        logging.error(f"Error deleting prompt: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500

# Call this function when the app starts
if __name__ == '__main__':
    logging.info("Starting the application")
    migrate_existing_users()  # Add this line
    app.run(debug=True)

